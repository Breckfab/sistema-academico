import os, io, json
from datetime import datetime, date
from functools import wraps
from flask import (Flask, render_template, redirect, url_for, flash,
                   request, session, jsonify, send_file, abort)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
#  APP CONFIG
# ─────────────────────────────────────────────
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'cambiar-en-produccion-xyz123')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///academia.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Mail config (Gmail por defecto; cambiar vars de entorno en Render)
app.config['MAIL_SERVER']   = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT']     = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS']  = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME', 'noreply@academia.com')

db    = SQLAlchemy(app)
mail  = Mail(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Debes iniciar sesión para acceder.'
s = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# ─────────────────────────────────────────────
#  MODELS
# ─────────────────────────────────────────────

class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuario'
    id           = db.Column(db.Integer, primary_key=True)
    username     = db.Column(db.String(80), unique=True, nullable=False)
    password_hash= db.Column(db.String(200), nullable=False)
    nombre       = db.Column(db.String(100))
    apellido     = db.Column(db.String(100))
    email        = db.Column(db.String(120), unique=True)
    rol          = db.Column(db.String(20), default='profesor')  # 'admin','profesor','demo'
    activo       = db.Column(db.Boolean, default=True)
    debe_cambiar_pass = db.Column(db.Boolean, default=False)
    materias     = db.relationship('Materia', back_populates='profesor', lazy='dynamic')

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)
    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)
    @property
    def nombre_completo(self):
        return f"{self.nombre or ''} {self.apellido or ''}".strip() or self.username


class CodigoRegistro(db.Model):
    __tablename__ = 'codigo_registro'
    id       = db.Column(db.Integer, primary_key=True)
    codigo   = db.Column(db.String(20), unique=True, nullable=False)
    usado    = db.Column(db.Boolean, default=False)
    creado_en= db.Column(db.DateTime, default=datetime.utcnow)


# Tabla de asociación alumno ↔ materia
inscripcion = db.Table('inscripcion',
    db.Column('alumno_id',  db.Integer, db.ForeignKey('alumno.id'),  primary_key=True),
    db.Column('materia_id', db.Integer, db.ForeignKey('materia.id'), primary_key=True),
    db.Column('anio_cursada', db.Integer),
    db.Column('fecha_inscripcion', db.Date, default=date.today)
)


class Materia(db.Model):
    __tablename__ = 'materia'
    id           = db.Column(db.Integer, primary_key=True)
    nombre       = db.Column(db.String(150), nullable=False)
    tipo         = db.Column(db.String(20), nullable=False)  # 'materia_anual','materia_cuatrimestral','taller'
    anio_academico = db.Column(db.Integer)
    horario      = db.Column(db.String(100))
    concepto     = db.Column(db.Text)
    programa     = db.Column(db.Text)
    notas_adicionales = db.Column(db.Text)
    # Semestres
    inicio_1er   = db.Column(db.Date)
    fin_1er      = db.Column(db.Date)
    inicio_2do   = db.Column(db.Date)
    fin_2do      = db.Column(db.Date)
    profesor_id  = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    profesor     = db.relationship('Usuario', back_populates='materias')
    alumnos      = db.relationship('Alumno', secondary=inscripcion,
                                   back_populates='materias', lazy='dynamic')
    calificaciones = db.relationship('Calificacion', back_populates='materia',
                                     cascade='all, delete-orphan')
    inasistencias  = db.relationship('Inasistencia', back_populates='materia',
                                     cascade='all, delete-orphan')


class Alumno(db.Model):
    __tablename__ = 'alumno'
    id       = db.Column(db.Integer, primary_key=True)
    nombre   = db.Column(db.String(100), nullable=False)
    apellido = db.Column(db.String(100), nullable=False)
    dni      = db.Column(db.String(20))
    materias = db.relationship('Materia', secondary=inscripcion,
                               back_populates='alumnos', lazy='dynamic')
    calificaciones = db.relationship('Calificacion', back_populates='alumno',
                                     cascade='all, delete-orphan')
    inasistencias  = db.relationship('Inasistencia', back_populates='alumno',
                                     cascade='all, delete-orphan')

    @property
    def nombre_completo(self):
        return f"{self.apellido}, {self.nombre}"


class Calificacion(db.Model):
    __tablename__ = 'calificacion'
    id          = db.Column(db.Integer, primary_key=True)
    alumno_id   = db.Column(db.Integer, db.ForeignKey('alumno.id'))
    materia_id  = db.Column(db.Integer, db.ForeignKey('materia.id'))
    # TPs almacenados como JSON: [{"titulo":"TP1","nota":8.5,"fecha":"2025-04-10","comentario":""}]
    tps_json    = db.Column(db.Text, default='[]')
    # Parcial
    parcial_nota    = db.Column(db.Float)
    parcial_fecha   = db.Column(db.Date)
    parcial_comentario = db.Column(db.Text)
    # R1
    r1_nota     = db.Column(db.Float)
    r1_fecha    = db.Column(db.Date)
    r1_comentario = db.Column(db.Text)
    # R2
    r2_nota     = db.Column(db.Float)
    r2_fecha    = db.Column(db.Date)
    r2_comentario = db.Column(db.Text)
    # Final
    final_nota  = db.Column(db.Float)
    final_fecha = db.Column(db.Date)
    final_comentario = db.Column(db.Text)
    final_forzado = db.Column(db.Boolean, default=False)
    # Concepto
    concepto_valor = db.Column(db.String(20))  # malo/bueno/muy bueno/excelente
    concepto_texto = db.Column(db.Text)
    alumno  = db.relationship('Alumno', back_populates='calificaciones')
    materia = db.relationship('Materia', back_populates='calificaciones')

    @property
    def tps(self):
        try: return json.loads(self.tps_json or '[]')
        except: return []

    @property
    def promedio_tps(self):
        tps = self.tps
        if not tps: return None
        notas = [t['nota'] for t in tps if t.get('nota') is not None]
        return round(sum(notas)/len(notas), 2) if notas else None

    @property
    def aprobo_parcial(self):
        """True si aprobó por alguna instancia (parcial, R1 o R2)."""
        if self.parcial_nota is not None and self.parcial_nota >= 4: return True
        if self.r1_nota is not None and self.r1_nota >= 4: return True
        if self.r2_nota is not None and self.r2_nota >= 4: return True
        return False


class Inasistencia(db.Model):
    __tablename__ = 'inasistencia'
    id          = db.Column(db.Integer, primary_key=True)
    alumno_id   = db.Column(db.Integer, db.ForeignKey('alumno.id'))
    materia_id  = db.Column(db.Integer, db.ForeignKey('materia.id'))
    cuatrimestre= db.Column(db.Integer)   # 1 o 2
    cantidad    = db.Column(db.Integer, default=0)
    alumno  = db.relationship('Alumno', back_populates='inasistencias')
    materia = db.relationship('Materia', back_populates='inasistencias')


class Reincorporatorio(db.Model):
    __tablename__ = 'reincorporatorio'
    id         = db.Column(db.Integer, primary_key=True)
    alumno_id  = db.Column(db.Integer, db.ForeignKey('alumno.id'))
    materia_id = db.Column(db.Integer, db.ForeignKey('materia.id'))
    nota       = db.Column(db.Float)
    fecha      = db.Column(db.Date)
    comentario = db.Column(db.Text)
    alumno     = db.relationship('Alumno')
    materia    = db.relationship('Materia')


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

@login_manager.user_loader
def load_user(uid):
    return db.session.get(Usuario, int(uid))


def is_demo():
    return current_user.is_authenticated and current_user.rol == 'demo'


def demo_block(f):
    """Decorator: bloquea escritura para usuario demo."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if is_demo():
            flash('Modo DEMO: los cambios no se guardan.', 'warning')
            return redirect(request.referrer or url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


def get_calificacion(alumno_id, materia_id):
    cal = Calificacion.query.filter_by(alumno_id=alumno_id, materia_id=materia_id).first()
    if not cal:
        cal = Calificacion(alumno_id=alumno_id, materia_id=materia_id)
        db.session.add(cal)
        db.session.flush()
    return cal


# ─────────────────────────────────────────────
#  AUTH ROUTES
# ─────────────────────────────────────────────

@app.route('/')
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user = Usuario.query.filter_by(username=username).first()
        if user and user.check_password(password) and user.activo:
            login_user(user)
            if user.debe_cambiar_pass:
                flash('Por favor cambiá tu contraseña provisoria.', 'info')
                return redirect(url_for('cambiar_password'))
            return redirect(url_for('dashboard'))
        flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        codigo   = request.form.get('codigo', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        nombre   = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        email    = request.form.get('email', '').strip()

        reg = CodigoRegistro.query.filter_by(codigo=codigo, usado=False).first()
        if not reg:
            flash('Error de código. Comuníquese con el administrador.', 'danger')
            return render_template('registro.html')

        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe.', 'danger')
            return render_template('registro.html')
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado.', 'danger')
            return render_template('registro.html')

        u = Usuario(username=username, nombre=nombre, apellido=apellido,
                    email=email, rol='profesor')
        u.set_password(password)
        reg.usado = True
        db.session.add(u)
        db.session.commit()
        flash('Registro exitoso. Podés iniciar sesión.', 'success')
        return redirect(url_for('login'))
    return render_template('registro.html')


@app.route('/recuperar_password', methods=['GET', 'POST'])
def recuperar_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        user  = Usuario.query.filter_by(email=email).first()
        if user:
            token = s.dumps(email, salt='reset-password')
            link  = url_for('reset_password', token=token, _external=True)
            try:
                msg = Message('Recuperar contraseña - Sistema Académico',
                              recipients=[email])
                msg.body = f'Para resetear tu contraseña hacé clic en:\n{link}\n\nEl enlace expira en 1 hora.'
                mail.send(msg)
            except Exception as e:
                app.logger.error(f'Mail error: {e}')
        flash('Si el email existe en el sistema, recibirás un enlace para resetear tu contraseña.', 'info')
        return redirect(url_for('login'))
    return render_template('recuperar_password.html')


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = s.loads(token, salt='reset-password', max_age=3600)
    except Exception:
        flash('El enlace expiró o es inválido.', 'danger')
        return redirect(url_for('login'))
    user = Usuario.query.filter_by(email=email).first_or_404()
    if request.method == 'POST':
        pw = request.form.get('password', '').strip()
        user.set_password(pw)
        user.debe_cambiar_pass = False
        db.session.commit()
        flash('Contraseña actualizada.', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html')


@app.route('/cambiar_password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if is_demo():
        flash('El usuario demo no puede cambiar contraseña.', 'warning')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        actual  = request.form.get('actual', '').strip()
        nueva   = request.form.get('nueva', '').strip()
        if not current_user.check_password(actual):
            flash('La contraseña actual es incorrecta.', 'danger')
        else:
            current_user.set_password(nueva)
            current_user.debe_cambiar_pass = False
            db.session.commit()
            flash('Contraseña cambiada correctamente.', 'success')
            return redirect(url_for('dashboard'))
    return render_template('cambiar_password.html')


# ─────────────────────────────────────────────
#  DASHBOARD
# ─────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.rol == 'admin':
        profesores = Usuario.query.filter(Usuario.rol.in_(['profesor','demo'])).all()
        materias   = Materia.query.all()
        alumnos    = Alumno.query.all()
        return render_template('dashboard_admin.html',
                               profesores=profesores, materias=materias, alumnos=alumnos)
    else:
        materias = Materia.query.filter_by(profesor_id=current_user.id).all()
        return render_template('dashboard_profe.html', materias=materias)


# ─────────────────────────────────────────────
#  MATERIAS
# ─────────────────────────────────────────────

@app.route('/materias')
@login_required
def listar_materias():
    if current_user.rol == 'admin':
        profe_id = request.args.get('profesor_id', type=int)
        if profe_id:
            materias = Materia.query.filter_by(profesor_id=profe_id).all()
        else:
            materias = Materia.query.all()
        profesores = Usuario.query.filter(Usuario.rol.in_(['profesor','demo'])).all()
        return render_template('materias.html', materias=materias,
                               profesores=profesores, profe_id=profe_id)
    else:
        materias = Materia.query.filter_by(profesor_id=current_user.id).all()
        return render_template('materias.html', materias=materias, profesores=None)


@app.route('/materias/nueva', methods=['GET', 'POST'])
@login_required
@demo_block
def nueva_materia():
    if request.method == 'POST':
        m = Materia(
            nombre    = request.form.get('nombre'),
            tipo      = request.form.get('tipo'),
            anio_academico = request.form.get('anio_academico', type=int),
            horario   = request.form.get('horario'),
            concepto  = request.form.get('concepto'),
            programa  = request.form.get('programa'),
            notas_adicionales = request.form.get('notas_adicionales'),
            inicio_1er= _parse_date(request.form.get('inicio_1er')),
            fin_1er   = _parse_date(request.form.get('fin_1er')),
            inicio_2do= _parse_date(request.form.get('inicio_2do')),
            fin_2do   = _parse_date(request.form.get('fin_2do')),
            profesor_id = current_user.id if current_user.rol != 'admin'
                          else request.form.get('profesor_id', current_user.id, type=int)
        )
        db.session.add(m)
        db.session.commit()
        flash('Materia creada correctamente.', 'success')
        return redirect(url_for('listar_materias'))
    profesores = Usuario.query.filter(Usuario.rol == 'profesor').all() \
                 if current_user.rol == 'admin' else None
    return render_template('nueva_materia.html', profesores=profesores)


@app.route('/materias/<int:mid>/editar', methods=['GET', 'POST'])
@login_required
@demo_block
def editar_materia(mid):
    m = Materia.query.get_or_404(mid)
    if current_user.rol != 'admin' and m.profesor_id != current_user.id:
        abort(403)
    if request.method == 'POST':
        m.nombre    = request.form.get('nombre')
        m.tipo      = request.form.get('tipo')
        m.anio_academico = request.form.get('anio_academico', type=int)
        m.horario   = request.form.get('horario')
        m.concepto  = request.form.get('concepto')
        m.programa  = request.form.get('programa')
        m.notas_adicionales = request.form.get('notas_adicionales')
        m.inicio_1er= _parse_date(request.form.get('inicio_1er'))
        m.fin_1er   = _parse_date(request.form.get('fin_1er'))
        m.inicio_2do= _parse_date(request.form.get('inicio_2do'))
        m.fin_2do   = _parse_date(request.form.get('fin_2do'))
        db.session.commit()
        flash('Materia actualizada.', 'success')
        return redirect(url_for('listar_materias'))
    return render_template('nueva_materia.html', materia=m, editando=True)


@app.route('/materias/<int:mid>/borrar', methods=['POST'])
@login_required
@demo_block
def borrar_materia(mid):
    m = Materia.query.get_or_404(mid)
    if current_user.rol != 'admin' and m.profesor_id != current_user.id:
        abort(403)
    db.session.delete(m)
    db.session.commit()
    flash('Materia eliminada.', 'success')
    return redirect(url_for('listar_materias'))


@app.route('/materias/<int:mid>/alumnos')
@login_required
def alumnos_por_materia(mid):
    m = Materia.query.get_or_404(mid)
    if current_user.rol != 'admin' and m.profesor_id != current_user.id:
        abort(403)
    anio = request.args.get('anio', type=int)
    query = db.session.query(Alumno, inscripcion.c.anio_cursada)\
              .join(inscripcion, Alumno.id == inscripcion.c.alumno_id)\
              .filter(inscripcion.c.materia_id == mid)
    if anio:
        query = query.filter(inscripcion.c.anio_cursada == anio)
    resultados = query.all()
    return render_template('alumnos_materia.html', materia=m, resultados=resultados, anio=anio)


# ─────────────────────────────────────────────
#  ALUMNOS
# ─────────────────────────────────────────────

@app.route('/alumnos')
@login_required
def listar_alumnos():
    q        = request.args.get('q', '').strip()
    materia_id = request.args.get('materia_id', type=int)
    anio     = request.args.get('anio', type=int)
    nota_min = request.args.get('nota_min', type=float)

    if current_user.rol == 'admin':
        base = Alumno.query
    else:
        ids = [m.id for m in Materia.query.filter_by(profesor_id=current_user.id)]
        base = Alumno.query.join(inscripcion).filter(inscripcion.c.materia_id.in_(ids))

    if q:
        base = base.filter(
            db.or_(Alumno.apellido.ilike(f'%{q}%'),
                   Alumno.nombre.ilike(f'%{q}%'),
                   Alumno.dni.ilike(f'%{q}%')))
    if materia_id:
        base = base.join(inscripcion, Alumno.id == inscripcion.c.alumno_id)\
                   .filter(inscripcion.c.materia_id == materia_id)
    if anio:
        base = base.join(inscripcion, Alumno.id == inscripcion.c.alumno_id)\
                   .filter(inscripcion.c.anio_cursada == anio)

    alumnos = base.distinct().all()

    mis_materias = Materia.query.filter_by(profesor_id=current_user.id).all() \
                   if current_user.rol != 'admin' \
                   else Materia.query.all()
    return render_template('alumnos.html', alumnos=alumnos, materias=mis_materias,
                           q=q, materia_id=materia_id, anio=anio)


@app.route('/alumnos/nuevo', methods=['GET', 'POST'])
@login_required
@demo_block
def nuevo_alumno():
    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        dni      = request.form.get('dni', '').strip()
        materia_ids   = request.form.getlist('materia_ids', type=int)
        anio_cursada  = request.form.get('anio_cursada', type=int) or date.today().year
        fecha_insc    = _parse_date(request.form.get('fecha_inscripcion')) or date.today()

        alumno = Alumno(nombre=nombre, apellido=apellido, dni=dni or None)
        db.session.add(alumno)
        db.session.flush()

        for mid in materia_ids:
            db.session.execute(inscripcion.insert().values(
                alumno_id=alumno.id, materia_id=mid,
                anio_cursada=anio_cursada, fecha_inscripcion=fecha_insc))

        db.session.commit()
        flash('Alumno registrado correctamente.', 'success')
        return redirect(url_for('listar_alumnos'))

    mis_materias = Materia.query.filter_by(profesor_id=current_user.id).all() \
                   if current_user.rol != 'admin' else Materia.query.all()
    return render_template('nuevo_alumno.html', materias=mis_materias,
                           hoy=date.today().isoformat())


@app.route('/alumnos/<int:aid>/editar', methods=['GET', 'POST'])
@login_required
@demo_block
def editar_alumno(aid):
    alumno = Alumno.query.get_or_404(aid)
    if request.method == 'POST':
        alumno.nombre   = request.form.get('nombre', '').strip()
        alumno.apellido = request.form.get('apellido', '').strip()
        alumno.dni      = request.form.get('dni', '').strip() or None
        db.session.commit()
        flash('Alumno actualizado.', 'success')
        return redirect(url_for('listar_alumnos'))
    return render_template('editar_alumno.html', alumno=alumno)


@app.route('/alumnos/<int:aid>/inscribir', methods=['GET', 'POST'])
@login_required
@demo_block
def inscribir_alumno(aid):
    """Inscribir un alumno existente a más materias."""
    alumno = Alumno.query.get_or_404(aid)
    if request.method == 'POST':
        materia_ids  = request.form.getlist('materia_ids', type=int)
        anio_cursada = request.form.get('anio_cursada', type=int) or date.today().year
        fecha_insc   = _parse_date(request.form.get('fecha_inscripcion')) or date.today()
        ya_inscritas = {m.id for m in alumno.materias}
        for mid in materia_ids:
            if mid not in ya_inscritas:
                db.session.execute(inscripcion.insert().values(
                    alumno_id=aid, materia_id=mid,
                    anio_cursada=anio_cursada, fecha_inscripcion=fecha_insc))
        db.session.commit()
        flash('Inscripción actualizada.', 'success')
        return redirect(url_for('listar_alumnos'))
    mis_materias = Materia.query.filter_by(profesor_id=current_user.id).all() \
                   if current_user.rol != 'admin' else Materia.query.all()
    return render_template('inscribir_alumno.html', alumno=alumno,
                           materias=mis_materias, hoy=date.today().isoformat())


@app.route('/alumnos/<int:aid>/borrar', methods=['POST'])
@login_required
@demo_block
def borrar_alumno(aid):
    alumno = Alumno.query.get_or_404(aid)
    db.session.delete(alumno)
    db.session.commit()
    flash('Alumno eliminado.', 'success')
    return redirect(url_for('listar_alumnos'))


# ─────────────────────────────────────────────
#  CALIFICACIONES
# ─────────────────────────────────────────────

@app.route('/calificaciones')
@login_required
def listar_calificaciones():
    q          = request.args.get('q', '').strip()
    materia_id = request.args.get('materia_id', type=int)
    anio       = request.args.get('anio', type=int)

    if current_user.rol == 'admin':
        mis_ids = [m.id for m in Materia.query.all()]
    else:
        mis_ids = [m.id for m in Materia.query.filter_by(profesor_id=current_user.id)]

    base = Calificacion.query.filter(Calificacion.materia_id.in_(mis_ids))

    if materia_id:
        base = base.filter_by(materia_id=materia_id)
    if q:
        base = base.join(Alumno).filter(
            db.or_(Alumno.apellido.ilike(f'%{q}%'),
                   Alumno.nombre.ilike(f'%{q}%')))
    if anio:
        base = base.join(Materia).filter(Materia.anio_academico == anio)

    cals = base.all()
    mis_materias = Materia.query.filter(Materia.id.in_(mis_ids)).all()
    return render_template('calificaciones.html', calificaciones=cals,
                           materias=mis_materias, q=q, materia_id=materia_id, anio=anio)


@app.route('/calificaciones/<int:aid>/<int:mid>', methods=['GET', 'POST'])
@login_required
@demo_block
def editar_calificacion(aid, mid):
    materia = Materia.query.get_or_404(mid)
    if current_user.rol != 'admin' and materia.profesor_id != current_user.id:
        abort(403)
    alumno = Alumno.query.get_or_404(aid)
    cal    = get_calificacion(aid, mid)

    if request.method == 'POST':
        action = request.form.get('action', 'save')
        if action == 'save':
            # TPs
            tp_titulos  = request.form.getlist('tp_titulo')
            tp_notas    = request.form.getlist('tp_nota')
            tp_fechas   = request.form.getlist('tp_fecha')
            tp_comentarios = request.form.getlist('tp_comentario')
            tps = []
            for t, n, f, c in zip(tp_titulos, tp_notas, tp_fechas, tp_comentarios):
                try: nota = float(n) if n.strip() else None
                except: nota = None
                tps.append({'titulo': t, 'nota': nota,
                            'fecha': f, 'comentario': c})
            cal.tps_json = json.dumps(tps)

            # Parcial
            cal.parcial_nota     = _to_float(request.form.get('parcial_nota'))
            cal.parcial_fecha    = _parse_date(request.form.get('parcial_fecha'))
            cal.parcial_comentario = request.form.get('parcial_comentario')
            # R1
            cal.r1_nota  = _to_float(request.form.get('r1_nota'))
            cal.r1_fecha = _parse_date(request.form.get('r1_fecha'))
            cal.r1_comentario = request.form.get('r1_comentario')
            # R2
            cal.r2_nota  = _to_float(request.form.get('r2_nota'))
            cal.r2_fecha = _parse_date(request.form.get('r2_fecha'))
            cal.r2_comentario = request.form.get('r2_comentario')
            # Final
            forzar = request.form.get('forzar_final') == '1'
            if materia.tipo in ('materia_anual', 'materia_cuatrimestral'):
                if not cal.aprobo_parcial and not forzar:
                    flash('El alumno no aprobó parcial/recuperatorios. '
                          'Debe marcar "Forzar habilitación de final".', 'warning')
                    db.session.rollback()
                    return redirect(url_for('editar_calificacion', aid=aid, mid=mid))
                cal.final_forzado = forzar and not cal.aprobo_parcial
            cal.final_nota  = _to_float(request.form.get('final_nota'))
            cal.final_fecha = _parse_date(request.form.get('final_fecha'))
            cal.final_comentario = request.form.get('final_comentario')
            # Concepto
            cal.concepto_valor = request.form.get('concepto_valor')
            cal.concepto_texto = request.form.get('concepto_texto')
            db.session.commit()
            flash('Calificaciones guardadas.', 'success')
        return redirect(url_for('editar_calificacion', aid=aid, mid=mid))

    return render_template('editar_calificacion.html',
                           alumno=alumno, materia=materia, cal=cal)


# ─────────────────────────────────────────────
#  INASISTENCIAS
# ─────────────────────────────────────────────

@app.route('/inasistencias')
@login_required
def listar_inasistencias():
    q          = request.args.get('q', '').strip()
    materia_id = request.args.get('materia_id', type=int)

    if current_user.rol == 'admin':
        mis_ids = [m.id for m in Materia.query.all()]
    else:
        mis_ids = [m.id for m in Materia.query.filter_by(profesor_id=current_user.id)]

    base = Inasistencia.query.filter(Inasistencia.materia_id.in_(mis_ids))
    if materia_id:
        base = base.filter_by(materia_id=materia_id)
    if q:
        base = base.join(Alumno).filter(
            db.or_(Alumno.apellido.ilike(f'%{q}%'),
                   Alumno.nombre.ilike(f'%{q}%')))

    # Agrupar por alumno+materia
    registros = base.all()
    mis_materias = Materia.query.filter(Materia.id.in_(mis_ids)).all()
    return render_template('inasistencias.html', registros=registros,
                           materias=mis_materias, q=q, materia_id=materia_id)


@app.route('/inasistencias/editar/<int:aid>/<int:mid>', methods=['GET', 'POST'])
@login_required
@demo_block
def editar_inasistencias(aid, mid):
    materia = Materia.query.get_or_404(mid)
    if current_user.rol != 'admin' and materia.profesor_id != current_user.id:
        abort(403)
    alumno = Alumno.query.get_or_404(aid)

    inas1 = Inasistencia.query.filter_by(alumno_id=aid, materia_id=mid, cuatrimestre=1).first()
    inas2 = Inasistencia.query.filter_by(alumno_id=aid, materia_id=mid, cuatrimestre=2).first()
    if not inas1:
        inas1 = Inasistencia(alumno_id=aid, materia_id=mid, cuatrimestre=1, cantidad=0)
        db.session.add(inas1)
    if not inas2:
        inas2 = Inasistencia(alumno_id=aid, materia_id=mid, cuatrimestre=2, cantidad=0)
        db.session.add(inas2)
    db.session.flush()

    if request.method == 'POST':
        inas1.cantidad = int(request.form.get('cuatri1', 0) or 0)
        inas2.cantidad = int(request.form.get('cuatri2', 0) or 0)
        db.session.commit()
        flash('Inasistencias guardadas.', 'success')
        return redirect(url_for('listar_inasistencias'))
    return render_template('editar_inasistencias.html',
                           alumno=alumno, materia=materia, inas1=inas1, inas2=inas2)


# ─────────────────────────────────────────────
#  REINCORPORATORIOS
# ─────────────────────────────────────────────

@app.route('/reincorporatorios')
@login_required
def listar_reincorporatorios():
    if current_user.rol == 'admin':
        mis_ids = [m.id for m in Materia.query.all()]
    else:
        mis_ids = [m.id for m in Materia.query.filter_by(profesor_id=current_user.id)]
    regs = Reincorporatorio.query.filter(Reincorporatorio.materia_id.in_(mis_ids)).all()
    mis_materias = Materia.query.filter(Materia.id.in_(mis_ids)).all()
    alumnos_propios = Alumno.query.join(inscripcion)\
                      .filter(inscripcion.c.materia_id.in_(mis_ids)).distinct().all()
    return render_template('reincorporatorios.html', regs=regs,
                           materias=mis_materias, alumnos=alumnos_propios)


@app.route('/reincorporatorios/nuevo', methods=['POST'])
@login_required
@demo_block
def nuevo_reincorporatorio():
    r = Reincorporatorio(
        alumno_id  = request.form.get('alumno_id', type=int),
        materia_id = request.form.get('materia_id', type=int),
        nota       = _to_float(request.form.get('nota')),
        fecha      = _parse_date(request.form.get('fecha')),
        comentario = request.form.get('comentario')
    )
    db.session.add(r)
    db.session.commit()
    flash('Reincorporatorio registrado.', 'success')
    return redirect(url_for('listar_reincorporatorios'))


@app.route('/reincorporatorios/<int:rid>/borrar', methods=['POST'])
@login_required
@demo_block
def borrar_reincorporatorio(rid):
    r = Reincorporatorio.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    flash('Reincorporatorio eliminado.', 'success')
    return redirect(url_for('listar_reincorporatorios'))


# ─────────────────────────────────────────────
#  ADMIN – Usuarios / Códigos
# ─────────────────────────────────────────────

@app.route('/admin/profesores')
@login_required
@admin_required
def admin_profesores():
    profesores = Usuario.query.filter(Usuario.rol.in_(['profesor'])).all()
    codigos    = CodigoRegistro.query.order_by(CodigoRegistro.creado_en.desc()).all()
    return render_template('admin_profesores.html', profesores=profesores, codigos=codigos)


@app.route('/admin/codigos/nuevo', methods=['POST'])
@login_required
@admin_required
def nuevo_codigo():
    codigo = request.form.get('codigo', '').strip()
    if CodigoRegistro.query.filter_by(codigo=codigo).first():
        flash('Código ya existente.', 'danger')
    else:
        db.session.add(CodigoRegistro(codigo=codigo))
        db.session.commit()
        flash(f'Código {codigo} generado.', 'success')
    return redirect(url_for('admin_profesores'))


@app.route('/admin/profesores/nuevo', methods=['POST'])
@login_required
@admin_required
def admin_nuevo_profesor():
    """Admin crea profesor directamente con contraseña provisoria."""
    username = request.form.get('username', '').strip()
    nombre   = request.form.get('nombre', '').strip()
    apellido = request.form.get('apellido', '').strip()
    email    = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()

    if Usuario.query.filter_by(username=username).first():
        flash('El nombre de usuario ya existe.', 'danger')
        return redirect(url_for('admin_profesores'))

    u = Usuario(username=username, nombre=nombre, apellido=apellido,
                email=email, rol='profesor', debe_cambiar_pass=True)
    u.set_password(password)
    db.session.add(u)
    db.session.commit()
    flash(f'Profesor {nombre} {apellido} creado. Contraseña provisoria: {password}', 'success')
    return redirect(url_for('admin_profesores'))


@app.route('/admin/profesores/<int:uid>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_profesor(uid):
    u = Usuario.query.get_or_404(uid)
    u.activo = not u.activo
    db.session.commit()
    flash(f'Usuario {"activado" if u.activo else "desactivado"}.', 'info')
    return redirect(url_for('admin_profesores'))


# ─────────────────────────────────────────────
#  EXPORTAR
# ─────────────────────────────────────────────

@app.route('/exportar')
@login_required
def exportar():
    tipo = request.args.get('tipo', 'excel')
    materia_id = request.args.get('materia_id', type=int)

    if current_user.rol == 'admin':
        mis_ids = [m.id for m in Materia.query.all()]
    else:
        mis_ids = [m.id for m in Materia.query.filter_by(profesor_id=current_user.id)]

    if materia_id and materia_id in mis_ids:
        mis_ids = [materia_id]

    cals = Calificacion.query.filter(Calificacion.materia_id.in_(mis_ids)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Calificaciones'

    headers = ['Apellido', 'Nombre', 'DNI', 'Materia', 'Año', 'Prom.TPs',
               'Parcial', 'R1', 'R2', 'Final', 'Concepto', 'Profesor']
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(fill_type='solid', fgColor='003366')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center')

    red_font = Font(bold=True, color='FF0000')
    for row, cal in enumerate(cals, 2):
        vals = [
            cal.alumno.apellido, cal.alumno.nombre,
            cal.alumno.dni or '', cal.materia.nombre,
            cal.materia.anio_academico, cal.promedio_tps,
            cal.parcial_nota, cal.r1_nota, cal.r2_nota,
            cal.final_nota, cal.concepto_valor or '',
            cal.materia.profesor.nombre_completo if cal.materia.profesor else ''
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            if isinstance(v, (int, float)) and v is not None and v <= 3:
                c.font = red_font

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True,
                     download_name='calificaciones.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─────────────────────────────────────────────
#  UTILS
# ─────────────────────────────────────────────

def _parse_date(s):
    if not s: return None
    try: return datetime.strptime(s, '%Y-%m-%d').date()
    except: return None

def _to_float(s):
    if not s: return None
    try: return float(s)
    except: return None


# ─────────────────────────────────────────────
#  INIT DB + SEED
# ─────────────────────────────────────────────

def seed_data():
    """Crear usuarios base y datos demo."""
    # Super Admin
    if not Usuario.query.filter_by(username='admin').first():
        a = Usuario(username='admin', nombre='Administrador', rol='admin',
                    email='admin@academia.com')
        a.set_password('Seamist123')
        db.session.add(a)

    # Demo
    if not Usuario.query.filter_by(username='demo').first():
        d = Usuario(username='demo', nombre='Demo', apellido='Sistema',
                    rol='demo', email='demo@academia.com')
        d.set_password('demo')
        db.session.add(d)
        db.session.flush()

        # Materias demo
        m1 = Materia(nombre='Matemática Aplicada', tipo='materia_anual',
                     anio_academico=2025, horario='Lunes 18:00-20:00',
                     concepto='Álgebra y análisis', profesor_id=d.id,
                     inicio_1er=date(2025, 3, 1), fin_1er=date(2025, 7, 31),
                     inicio_2do=date(2025, 8, 1), fin_2do=date(2025, 12, 15))
        m2 = Materia(nombre='Diseño Web', tipo='taller',
                     anio_academico=2025, horario='Miércoles 16:00-18:00',
                     concepto='HTML, CSS, JS', profesor_id=d.id)
        m3 = Materia(nombre='Programación I', tipo='materia_cuatrimestral',
                     anio_academico=2025, horario='Jueves 19:00-21:00',
                     concepto='Lógica y algoritmos', profesor_id=d.id)
        db.session.add_all([m1, m2, m3])
        db.session.flush()

        # Alumnos demo
        alumnos_demo = [
            Alumno(nombre='Ana', apellido='García', dni='30111222'),
            Alumno(nombre='Carlos', apellido='López', dni='31222333'),
            Alumno(nombre='Lucía', apellido='Martínez', dni='32333444'),
        ]
        db.session.add_all(alumnos_demo)
        db.session.flush()

        for al in alumnos_demo:
            for m in [m1, m2, m3]:
                db.session.execute(inscripcion.insert().values(
                    alumno_id=al.id, materia_id=m.id,
                    anio_cursada=2025, fecha_inscripcion=date(2025, 3, 1)))
            db.session.flush()
            # Calificaciones demo
            for m in [m1, m2, m3]:
                cal = Calificacion(alumno_id=al.id, materia_id=m.id,
                    tps_json=json.dumps([
                        {'titulo':'TP1','nota':8.0,'fecha':'2025-04-10','comentario':'Buen trabajo'},
                        {'titulo':'TP2','nota':7.5,'fecha':'2025-05-15','comentario':'Correcto'},
                    ]),
                    parcial_nota=7.0, parcial_fecha=date(2025,6,10),
                    r1_nota=None, r2_nota=None,
                    final_nota=8.0 if m.tipo != 'taller' else None,
                    final_fecha=date(2025,12,5) if m.tipo != 'taller' else None,
                    concepto_valor='muy bueno')
                db.session.add(cal)

        db.session.commit()


def create_tables():
    db.create_all()
    seed_data()
    db.session.commit()


with app.app_context():
    create_tables()


if __name__ == '__main__':
    app.run(debug=True)
